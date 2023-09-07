import openpyxl
import os.path

@staticmethod
def getSheet(filepath="C:\\Workspace\\Python_Workspace\\com.orangehrmlive.management.hr.ui_automation_selframework\\testdata\\testData.xlsx"):
    global sheet
    try:
        if not os.path.isfile(filepath):
            raise FileNotFoundError(filepath + " does not exits")
        else:
            workbook = openpyxl.load_workbook()
            sheet = workbook["Sheet1"]
    except IndexError:
        print("Exception occured")
    return sheet

@staticmethod
def maxRowCount():
    sheet = getSheet()
    maxRow = int(sheet.max_row)
    assert maxRow > 0
    return maxRow

@staticmethod
def maxColCount():
    sheet = getSheet()
    maxCol = int(sheet.max_column)
    assert maxCol > 0
    return maxCol

@staticmethod
def minRowCount():
    sheet = getSheet()
    assert sheet.min_row > 0
    return sheet.min_row

@staticmethod
def minColCount():
    sheet = getSheet()
    assert sheet.min_column > 0
    return sheet.min_column

@staticmethod
def readCellData(rowNum:int, colNum:int):
    sheet = getSheet()
    return [sheet.cell(rowNum, colNum).value]

@staticmethod
def writeDataIntoCell(rowNum:int, colNum:int, data):
    workbook = openpyxl.load_workbook("C:\\Workspace\\Python_Workspace\\com.orangehrmlive.management.hr.ui_automation_selframework\\testdata\\testData.xlsx")
    # workbook = openpyxl.load_workbook(r"C:\Workspace\Python_Workspace\com.orangehrmlive.management.hr.ui_automation_selframework\testdata\testData.xlsx")
    sheet = workbook["Sheet1"]
    sheet.cell(rowNum, colNum).value = data
    workbook.save("C:\\Workspace\\Python_Workspace\\com.orangehrmlive.management.hr.ui_automation_selframework\\testdata\\testData.xlsx")


# self.logger.info(Exl.maxRowCount())
# self.logger.info(Exl.maxColCount())
# self.logger.info(Exl.minRowCount())
# self.logger.info(Exl.minColCount())
# for i in range(Exl.minRowCount() + 1, Exl.maxRowCount() + 1):
#     for j in range(Exl.minColCount(), Exl.maxColCount() + 1):
#         self.logger.info("###########################")
#         self.logger.info(Exl.readCellData(i, j))